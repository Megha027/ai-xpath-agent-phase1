from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelWriter:

    def write(self, elements, file_name="reports/xpath_report.xlsx"):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "XPath Report"

        headers = [
            "Tag",
            "ID",
            "Name",
            "Text",
            "AI XPath",
            "Validation Status"
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Write data
        row = 2

        for element in elements:

            sheet.cell(row=row, column=1).value = element.tag
            sheet.cell(row=row, column=2).value = element.element_id
            sheet.cell(row=row, column=3).value = element.name
            sheet.cell(row=row, column=4).value = element.text
            sheet.cell(row=row, column=5).value = element.ai_xpath
            sheet.cell(row=row, column=6).value = element.validation_status

            row += 1

        workbook.save(file_name)

        print(f"\n✅ Excel report saved at : {file_name}")